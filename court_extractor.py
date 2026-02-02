#!/usr/bin/env python3
"""
Court Decision Extractor
Extracts structured information from court decisions using AI and outputs to Excel.
"""

import os
import sys
import json
from pathlib import Path
from typing import Dict, List, Optional
from datetime import datetime

# Document parsing
import PyPDF2
from docx import Document

# AI extraction
from openai import OpenAI

# Excel generation
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class CourtDecisionExtractor:
    """Extract structured information from court decisions."""
    
    # Extraction schema
    FIELDS = [
        {"name": "case_name", "label": "Case Name", "type": "free_response"},
        {"name": "court", "label": "Court", "type": "free_response"},
        {"name": "judges", "label": "Judge(s)", "type": "free_response"},
        {"name": "court_type", "label": "Court Type", "type": "classify", 
         "options": ["State", "Federal"]},
        {"name": "parties", "label": "Parties", "type": "free_response"},
        {"name": "counsel", "label": "Counsel", "type": "free_response"},
        {"name": "subject_matter", "label": "Subject Matter", "type": "classify",
         "options": ["Intellectual Property", "Contract", "Tort", "Civil Rights", 
                    "Constitutional", "Antitrust", "Criminal", "State Regulatory", 
                    "Federal Regulatory", "Other"]},
        {"name": "procedural_history", "label": "Procedural History", "type": "free_response"},
        {"name": "procedural_posture", "label": "Procedural Posture", "type": "free_response"},
        {"name": "claims", "label": "Claim(s)", "type": "free_response"},
        {"name": "holdings", "label": "Holding(s)", "type": "free_response"},
        {"name": "statutes_cited", "label": "Statutes Cited", "type": "free_response"},
        {"name": "cases_cited", "label": "Cases Cited", "type": "free_response"},
        {"name": "order", "label": "Order", "type": "free_response"},
        {"name": "concurrence", "label": "Concurrence", "type": "free_response"},
        {"name": "dissent", "label": "Dissent", "type": "free_response"},
    ]
    
    def __init__(self):
        """Initialize the extractor with OpenAI client."""
        self.client = OpenAI()
        
    def parse_document(self, file_path: str) -> str:
        """
        Parse document and extract text content.
        
        Args:
            file_path: Path to the court decision document
            
        Returns:
            Extracted text content
        """
        path = Path(file_path)
        
        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        
        # Handle different file types
        if path.suffix.lower() == '.pdf':
            return self._parse_pdf(file_path)
        elif path.suffix.lower() in ['.docx', '.doc']:
            return self._parse_docx(file_path)
        elif path.suffix.lower() == '.txt':
            return self._parse_txt(file_path)
        else:
            raise ValueError(f"Unsupported file type: {path.suffix}")
    
    def _parse_pdf(self, file_path: str) -> str:
        """Extract text from PDF file."""
        text = []
        with open(file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            for page in pdf_reader.pages:
                text.append(page.extract_text())
        return '\n'.join(text)
    
    def _parse_docx(self, file_path: str) -> str:
        """Extract text from DOCX file."""
        doc = Document(file_path)
        return '\n'.join([para.text for para in doc.paragraphs])
    
    def _parse_txt(self, file_path: str) -> str:
        """Extract text from TXT file."""
        with open(file_path, 'r', encoding='utf-8') as file:
            return file.read()
    
    def extract_information(self, text: str, filename: str = "") -> Dict:
        """
        Extract structured information from court decision text using AI.
        
        Args:
            text: Court decision text
            filename: Original filename for reference
            
        Returns:
            Dictionary with extracted information
        """
        # Build extraction prompt
        prompt = self._build_extraction_prompt(text)
        
        print(f"Extracting information from {filename}...")
        
        try:
            # Call OpenAI API with structured output
            response = self.client.chat.completions.create(
                model="gpt-4.1-mini",
                messages=[
                    {"role": "system", "content": "You are a legal document analyst specializing in extracting structured information from court decisions."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.1,
                max_tokens=4000
            )
            
            # Parse JSON response
            result = json.loads(response.choices[0].message.content)
            result['source_file'] = filename
            result['extraction_date'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            return result
            
        except Exception as e:
            print(f"Error during extraction: {e}")
            # Return empty structure on error
            return self._empty_result(filename)
    
    def _build_extraction_prompt(self, text: str) -> str:
        """Build the extraction prompt for the AI."""
        
        # Truncate text if too long (keep first 30k chars)
        if len(text) > 30000:
            text = text[:30000] + "\n\n[Document truncated for processing]"
        
        prompt = f"""Extract the following information from this court decision and return it as a JSON object.

For each field:
- If the information is clearly present, extract it accurately
- If the information is not found or not applicable, use "N/A"
- For classification fields, choose the most appropriate option from the provided list
- For free response fields, provide concise but complete answers

COURT DECISION TEXT:
{text}

EXTRACTION SCHEMA:

"""
        
        # Add field descriptions
        for field in self.FIELDS:
            prompt += f"\n{field['label']} ({field['name']}):\n"
            if field['type'] == 'classify':
                prompt += f"  Type: Classification - Choose ONE from: {', '.join(field['options'])}\n"
            else:
                prompt += f"  Type: Free response - Extract relevant information\n"
        
        prompt += """

Return ONLY a valid JSON object with this exact structure:
{
"""
        for i, field in enumerate(self.FIELDS):
            comma = "," if i < len(self.FIELDS) - 1 else ""
            prompt += f'  "{field["name"]}": "extracted value here"{comma}\n'
        
        prompt += """}

Do not include any explanation or text outside the JSON object."""
        
        return prompt
    
    def _empty_result(self, filename: str) -> Dict:
        """Create an empty result structure."""
        result = {field['name']: 'N/A' for field in self.FIELDS}
        result['source_file'] = filename
        result['extraction_date'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        return result
    
    def process_documents(self, file_paths: List[str]) -> List[Dict]:
        """
        Process multiple court decision documents.
        
        Args:
            file_paths: List of file paths to process
            
        Returns:
            List of extraction results
        """
        results = []
        
        for i, file_path in enumerate(file_paths, 1):
            print(f"\n[{i}/{len(file_paths)}] Processing: {file_path}")
            
            try:
                # Parse document
                text = self.parse_document(file_path)
                
                # Extract information
                result = self.extract_information(text, Path(file_path).name)
                results.append(result)
                
                print(f"✓ Successfully extracted information")
                
            except Exception as e:
                print(f"✗ Error processing {file_path}: {e}")
                results.append(self._empty_result(Path(file_path).name))
        
        return results
    
    def export_to_excel(self, results: List[Dict], output_path: str):
        """
        Export extraction results to a professional Excel file.
        
        Args:
            results: List of extraction results
            output_path: Path for output Excel file
        """
        print(f"\nGenerating Excel file: {output_path}")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Court Decisions"
        
        # Theme colors (Elegant Black)
        THEME = {
            'primary': '2D2D2D',
            'light': 'E5E5E5',
            'accent': '2D2D2D',
            'text': '000000'
        }
        
        # Hide gridlines
        ws.sheet_view.showGridLines = False
        
        # Set column A as left margin
        ws.column_dimensions['A'].width = 3
        
        # Title
        ws['B2'] = "Court Decision Extraction Results"
        ws['B2'].font = Font(name='Georgia', size=18, bold=True, color=THEME['primary'])
        
        # Metadata
        ws['B3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['B3'].font = Font(name='Calibri', size=10, color='666666')
        
        ws['B4'] = f"Total Cases: {len(results)}"
        ws['B4'].font = Font(name='Calibri', size=10, color='666666')
        
        # Table starts at row 6
        header_row = 6
        data_start_row = header_row + 1
        
        # Headers
        headers = ['Source File', 'Extraction Date'] + [field['label'] for field in self.FIELDS]
        
        for col_idx, header in enumerate(headers, start=2):  # Start from column B
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=THEME['primary'], end_color=THEME['primary'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(
                bottom=Side(style='thin', color='000000')
            )
        
        # Data rows
        for row_idx, result in enumerate(results, start=data_start_row):
            # Source file and extraction date
            ws.cell(row=row_idx, column=2, value=result.get('source_file', 'N/A'))
            ws.cell(row=row_idx, column=3, value=result.get('extraction_date', 'N/A'))
            
            # Extracted fields
            for col_idx, field in enumerate(self.FIELDS, start=4):
                value = result.get(field['name'], 'N/A')
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                
                # Alternate row colors
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color='F9F9F9', end_color='F9F9F9', fill_type='solid')
        
        # Set column widths
        column_widths = {
            'B': 25,  # Source File
            'C': 18,  # Extraction Date
            'D': 30,  # Case Name
            'E': 25,  # Court
            'F': 20,  # Judge(s)
            'G': 12,  # Court Type
            'H': 25,  # Parties
            'I': 25,  # Counsel
            'J': 18,  # Subject Matter
            'K': 35,  # Procedural History
            'L': 30,  # Procedural Posture
            'M': 35,  # Claim(s)
            'N': 35,  # Holding(s)
            'O': 30,  # Statutes Cited
            'P': 30,  # Cases Cited
            'Q': 30,  # Order
            'R': 30,  # Concurrence
            'S': 30,  # Dissent
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Set row heights
        ws.row_dimensions[header_row].height = 40
        for row_idx in range(data_start_row, data_start_row + len(results)):
            ws.row_dimensions[row_idx].height = 60
        
        # Freeze panes (freeze header row)
        ws.freeze_panes = f'B{data_start_row}'
        
        # Add auto-filter
        last_col = get_column_letter(len(headers) + 1)
        ws.auto_filter.ref = f'B{header_row}:{last_col}{data_start_row + len(results) - 1}'
        
        # Save workbook
        wb.save(output_path)
        print(f"✓ Excel file created successfully: {output_path}")


def main():
    """Main entry point for the program."""
    
    print("=" * 70)
    print("Court Decision Extractor")
    print("=" * 70)
    
    # Check for input files
    if len(sys.argv) < 2:
        print("\nUsage: python court_extractor.py <file1> [file2] [file3] ...")
        print("\nSupported formats: PDF, DOCX, TXT")
        print("\nExample:")
        print("  python court_extractor.py decision1.pdf decision2.pdf")
        sys.exit(1)
    
    # Get input files
    input_files = sys.argv[1:]
    
    # Validate files exist
    valid_files = []
    for file_path in input_files:
        if Path(file_path).exists():
            valid_files.append(file_path)
        else:
            print(f"Warning: File not found: {file_path}")
    
    if not valid_files:
        print("\nError: No valid input files found.")
        sys.exit(1)
    
    print(f"\nFound {len(valid_files)} file(s) to process\n")
    
    # Initialize extractor
    extractor = CourtDecisionExtractor()
    
    # Process documents
    results = extractor.process_documents(valid_files)
    
    # Generate output filename
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = f"court_decisions_extracted_{timestamp}.xlsx"
    
    # Export to Excel
    extractor.export_to_excel(results, output_file)
    
    print("\n" + "=" * 70)
    print("Extraction Complete!")
    print(f"Results saved to: {output_file}")
    print("=" * 70)


if __name__ == "__main__":
    main()
